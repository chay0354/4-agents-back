from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import Optional, List, Dict
import os
import json
from datetime import datetime, timezone
from dotenv import load_dotenv
from agents.workflow import AgentWorkflow
from database.mongodb import MongoDBClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

load_dotenv()

app = FastAPI(title="4-Agents MOP System", version="1.0.0")

# CORS middleware - allow all origins
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,  # Must be False when allow_origins=["*"]
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize MongoDB client
db_client = MongoDBClient()

# Kernel state - tracks if hard stop is pressed
kernel_hard_stop = False

# Current agent tracking - tracks which agent is currently running
current_agent: Optional[str] = None

# Stop history - tracks all stop events with timestamps
kernel_stop_history: List[Dict] = []

class ProblemRequest(BaseModel):
    problem: str

@app.get("/")
async def root():
    return {"message": "4-Agents MOP System API", "status": "running"}

@app.get("/health")
async def health():
    return {"status": "healthy", "database": db_client.is_connected()}

@app.get("/kernel")
async def kernel_check():
    """
    Kernel endpoint that decides if analysis should continue
    Returns 'ok' to continue, 'stop' to halt
    """
    global kernel_hard_stop
    if kernel_hard_stop:
        return {"status": "stop", "message": "Hard stop activated"}
    return {"status": "ok", "message": "Continue"}

@app.post("/kernel/stop")
async def kernel_stop():
    """
    Activate hard stop - prevents analysis from continuing
    """
    global kernel_hard_stop, kernel_stop_history, current_agent
    kernel_hard_stop = True
    # Record stop event in history with current agent info
    stop_event = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "stop",
        "stopped_agent": current_agent or "Unknown"
    }
    kernel_stop_history.append(stop_event)
    return {"status": "stopped", "message": "Hard stop activated"}

@app.post("/kernel/reset")
async def kernel_reset():
    """
    Reset hard stop - allows analysis to continue
    """
    global kernel_hard_stop, kernel_stop_history
    kernel_hard_stop = False
    # Record reset event in history
    reset_event = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "action": "reset",
        "status": "deactivated"
    }
    kernel_stop_history.append(reset_event)
    return {"status": "reset", "message": "Hard stop reset"}

@app.post("/analyze")
async def analyze_problem(request: ProblemRequest):
    """
    Main endpoint to analyze a problem using the 4-agent system
    Streams agent responses in real-time
    """
    async def generate():
        try:
            # Reset kernel state when starting new analysis
            global kernel_hard_stop, current_agent
            kernel_hard_stop = False
            current_agent = None
            
            workflow = AgentWorkflow(db_client)
            all_responses = {}
            final_kernel_decision = None  # Track final kernel decision
            
            async for update in workflow.process_problem_stream(request.problem):
                # Track current agent from updates and update stop history if stopped
                if update.get("agent") and update.get("agent") != "system":
                    if update.get("status") == "thinking":
                        current_agent = update.get("agent")
                    elif update.get("status") == "complete":
                        # Keep current agent until next one starts or analysis stops
                        current_agent = update.get("agent")
                    elif update.get("status") == "stopped" and update.get("stopped_agent"):
                        # Update the most recent stop event with the actual stopped agent
                        global kernel_stop_history
                        if kernel_stop_history and kernel_stop_history[-1].get("action") == "stop":
                            kernel_stop_history[-1]["stopped_agent"] = update.get("stopped_agent")
                
                # Track final kernel decision from updates
                if "kernel_decision" in update and update["kernel_decision"] is not None:
                    final_kernel_decision = update["kernel_decision"]
                
                # Collect all responses for final save
                if update.get("status") == "complete" and "response" in update:
                    all_responses[update["agent"]] = update["response"]
                
                # Stream the update immediately - each agent completes before next starts
                update_json = json.dumps(update)
                print(f"Streaming update: agent={update.get('agent')}, status={update.get('status')}, kernel_decision={update.get('kernel_decision')}")
                yield f"data: {update_json}\n\n"
            
            # Determine final kernel decision if not set (defaults to "N" if completed successfully)
            if final_kernel_decision is None:
                final_kernel_decision = "N"  # N = Normal completion (no hard stop occurred)
            
            # Save to database
            result = {
                "problem": request.problem,
                "responses": all_responses,
                "final_insights": all_responses.get("final", ""),
                "status": "completed" if final_kernel_decision == "N" else "stopped",
                "kernel_decision": final_kernel_decision,  # N = Normal, L = Limited/Stopped
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            db_client.save_analysis(result)
            
        except Exception as e:
            error_update = {
                "agent": "error",
                "status": "error",
                "message": str(e),
                "kernel_decision": None  # Error state
            }
            yield f"data: {json.dumps(error_update)}\n\n"
    
    return StreamingResponse(
        generate(), 
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Disable nginx buffering
        }
    )

@app.get("/analyses")
async def get_analyses():
    """
    Get all previous analyses
    """
    try:
        analyses = db_client.get_all_analyses()
        return {"analyses": analyses}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/analyses/{analysis_id}")
async def get_analysis(analysis_id: str):
    """
    Get a specific analysis by ID
    """
    try:
        analysis = db_client.get_analysis(analysis_id)
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")
        return analysis
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/kernel/history")
async def get_kernel_history():
    """
    Get history of all kernel stop/reset events
    """
    global kernel_stop_history
    return {"history": kernel_stop_history, "count": len(kernel_stop_history)}

@app.get("/kernel/history/export")
async def export_kernel_history():
    """
    Export kernel stop history as Excel file
    """
    global kernel_stop_history
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kernel Stop History"
    
    # Define header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = ["Timestamp", "Action", "Stopped Agent"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for row_num, event in enumerate(kernel_stop_history, 2):
        ws.cell(row=row_num, column=1, value=event.get("timestamp", ""))
        ws.cell(row=row_num, column=2, value=event.get("action", "").upper())
        # For stop events, show which agent was stopped; for reset events, show "N/A"
        if event.get("action") == "stop":
            stopped_agent = event.get("stopped_agent", "Unknown")
            # Format agent name nicely
            agent_name = stopped_agent.replace("_", " ").title() if stopped_agent != "Unknown" else "Unknown"
            ws.cell(row=row_num, column=3, value=agent_name)
        else:
            ws.cell(row=row_num, column=3, value="N/A")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with current date
    filename = f"kernel_stop_history_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

